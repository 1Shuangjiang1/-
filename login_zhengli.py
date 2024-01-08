from tkinter import  messagebox
from openpyxl import Workbook, load_workbook
import os
#本文件是登录界面

class loglogin:
    def login(self, username, password):
        if self.check_blacklist(username):
            messagebox.showerror("登录失败", "该用户已被列入黑名单，无法登录！")
            return False
        if not self.check_user_exists(username):
            messagebox.showerror("登录失败", "用户名不存在！")
            return False
        if self.check_password(username, password):
            messagebox.showinfo("登录成功", f"欢迎回来，{username}!")
            self.username = username  # 保存用户名
            # 邮箱已在 check_password 方法中设置
            messagebox.showinfo("登录信息", f"用户名: {self.username}, 邮箱: {self.email}")
            if self.auto_import_enabled:
                self.auto_import_contacts()  # 自动导入联系人
            return True  # 登录成功
        else:
            messagebox.showerror("登录失败", "密码错误！")
            self.add_to_blacklist(username)
            messagebox.showinfo("提示", "由于连续登录失败，该用户已被加入黑名单。")
            return False  # 登录失败


    def init_user_excel(self):
        if not os.path.exists(self.user_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["用户名", "密码", "邮箱"])  # 添加电子邮件列
            wb.save(self.user_file)
        else:
            wb = load_workbook(self.user_file)
        return wb


    def init_blacklist_excel(self):
        if not os.path.exists(self.blacklist_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Username"])
            wb.save(self.blacklist_file)
        else:
            wb = load_workbook(self.blacklist_file)
        return wb


    def check_user_exists(self, username):
        wb = self.init_user_excel()
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if username == row[0].value:
                return True
        return False


    def register(self, username, password, email):  # 添加电子邮件参数
        wb = self.init_user_excel()
        ws = wb.active
        ws.append([username, password, email])
        wb.save(self.user_file)


    def check_password(self, username, password):
        wb = self.init_user_excel()
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if username == row[0].value and password == row[1].value:
                # 设置电子邮件属性
                self.email = row[2].value
                return True
        return False


    def add_to_blacklist(self, username):
        wb = self.init_blacklist_excel()
        ws = wb.active
        ws.append([username])
        wb.save(self.blacklist_file)


    def check_blacklist(self, username):
        wb = self.init_blacklist_excel()
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if username == row[0].value:
                return True
        return False