import openpyxl
from openpyxl import Workbook
import os

# 文件路径
user_file = "users.xlsx"
blacklist_file = "blacklist.xlsx"

# 初始化或读取用户Excel文件
def init_user_excel():
    if not os.path.exists(user_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Username", "Password"])
        wb.save(user_file)
    else:
        wb = openpyxl.load_workbook(user_file)
    return wb

# 初始化或读取黑名单Excel文件
def init_blacklist_excel():
    if not os.path.exists(blacklist_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Username"])
        wb.save(blacklist_file)
    else:
        wb = openpyxl.load_workbook(blacklist_file)
    return wb

# 检查用户名是否已经存在
def check_user_exists(username):
    wb = init_user_excel()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if username == row[0].value:
            return True
    return False

# 添加新用户
def register(username, password):
    wb = init_user_excel()
    ws = wb.active
    ws.append([username, password])
    wb.save(user_file)

# 检查密码是否正确
def check_password(username, password):
    wb = init_user_excel()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if username == row[0].value and password == row[1].value:
            return True
    return False

# 添加用户到黑名单
def add_to_blacklist(username):
    wb = init_blacklist_excel()
    ws = wb.active
    ws.append([username])
    wb.save(blacklist_file)

# 检查用户是否在黑名单
def check_blacklist(username):
    wb = init_blacklist_excel()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if username == row[0].value:
            return True
    return False

# 用户注册
def user_register():
    username = input("请输入用户名: ")
    if check_user_exists(username):
        print("用户名已存在，请重新输入！")
    else:
        password = input("请输入密码: ")
        register(username, password)
        print("注册成功！")

# 用户登录
def user_login():
    username = input("请输入用户名: ")
    if check_blacklist(username):
        print("该用户已被列入黑名单，无法登录！")
        return
    if not check_user_exists(username):
        print("用户名不存在，请重新输入！")
        return
    password = input("请输入密码: ")
    if check_password(username, password):
        print("登录成功！")
    else:
        print("密码错误！")
        add_to_blacklist(username)
        print("由于连续登录失败，该用户已被加入黑名单。")

# 主程序
def main():
    while True:
        action = input("请选择操作：1 - 登录，2 - 注册，其他 - 退出：")
        if action == '1':
            user_login()
        elif action == '2':
            user_register()
        else:
            break

if __name__ == "__main__":
    main()
