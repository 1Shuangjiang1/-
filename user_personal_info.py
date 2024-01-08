from ContactBook_class import *
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
import tkinter as tk
from PIL import Image, ImageTk, ImageDraw
from openpyxl import load_workbook
import os
import datetime

def edit_user_info(self):
    edit_window = tk.Toplevel(self.root)
    edit_window.title("编辑个人信息")

    # 输入生日
    tk.Label(edit_window, text="生日 (YYYY-MM-DD):").pack()
    birthday_entry = tk.Entry(edit_window)
    birthday_entry.insert(0, self.user_info.get('birthday', ''))
    birthday_entry.pack()

    # 输入联系电话
    tk.Label(edit_window, text="联系电话:").pack()
    phone_entry = tk.Entry(edit_window)
    phone_entry.insert(0, self.user_info.get('phone', ''))
    phone_entry.pack()

    # 输入地址
    tk.Label(edit_window, text="地址:").pack()
    address_entry = tk.Entry(edit_window)
    address_entry.insert(0, self.user_info.get('address', ''))
    address_entry.pack()

    # 选择头像
    tk.Button(edit_window, text="选择头像", command=self.select_avatar).pack()

    # 保存按钮
    tk.Button(edit_window, text="保存",
              command=lambda: self.save_user_info(
                  birthday_entry.get(),
                  phone_entry.get(),
                  address_entry.get(),
                  edit_window)).pack()


def select_avatar(self):
    file_path = filedialog.askopenfilename(title="选择头像图片", filetypes=[("图片文件", "*.jpg *.png")])
    if file_path:
        self.user_info['avatar'] = file_path


def save_user_info(self, birthday, phone, address, window):
    self.user_info['birthday'] = birthday
    self.user_info['phone'] = phone
    self.user_info['address'] = address
    self.save_to_excel()
    self.save_avatar()
    messagebox.showinfo("保存成功", "个人信息已更新！")
    window.destroy()


def save_to_excel(self):
    # 文件名
    filename = "UserInfo.xlsx"
    # 检查文件是否存在
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # 添加表头
        sheet.append(["用户名", "生日", "联系电话", "地址"])

    # 检查是否已存在该用户信息
    user_row = None
    for row in sheet.iter_rows(min_row=2):  # 从第二行开始查找
        if row[0].value == self.username:  # 假设用户名存储在第一列
            user_row = row
            break

    # 更新或添加用户信息
    if user_row:  # 已存在该用户，更新信息
        user_row[1].value = self.user_info.get('birthday', '')
        user_row[2].value = self.user_info.get('phone', '')
        user_row[3].value = self.user_info.get('address', '')
    else:  # 不存在该用户，添加新行
        sheet.append([self.username, self.user_info.get('birthday', ''), self.user_info.get('phone', ''),
                      self.user_info.get('address', '')])

    # 保存工作簿
    workbook.save(filename)


def save_avatar(self):
    # 确保用户名和头像路径已经获取
    if self.username and self.user_info.get('avatar'):
        image_path = self.user_info['avatar']
        # 获取文件扩展名
        _, ext = os.path.splitext(image_path)
        # 定义新的文件名
        new_filename = f"{self.username}avatar{ext}"
        # 获取当前工作目录
        cwd = os.getcwd()
        new_file_path = os.path.join(cwd, new_filename)

        # 复制图片到当前目录下并重命名为用户名+avatar.ext
        try:
            shutil.copy(image_path, new_file_path)
            print(f"头像已保存: {new_file_path}")
        except Exception as e:
            print(f"保存头像失败: {e}")

def show_user_info(self):
    # 创建窗口
    info_window = tk.Toplevel(self.root)
    info_window.title(f"{self.username}的个人信息")

    # 读取用户信息
    self.read_user_info()

    # 显示信息
    tk.Label(info_window, text=f"用户名: {self.username}").pack()
    tk.Label(info_window, text=f"生日: {self.user_info.get('birthday', '未设置')}").pack()
    tk.Label(info_window, text=f"电话: {self.user_info.get('phone', '未设置')}").pack()
    tk.Label(info_window, text=f"地址: {self.user_info.get('address', '未设置')}").pack()

    # 处理头像
    # avatar_path = f"{self.username}avatar.jpg"  # 假设头像命名规则
    # if os.path.exists(avatar_path):
    #     print('exisit')
    #     image = Image.open(avatar_path)
    #     image = self.crop_to_circle(image)
    #     photo = ImageTk.PhotoImage(image)
    #     avatar_label = tk.Label(info_window, image=photo)
    #     avatar_label.image = photo  # keep a reference!
    #     avatar_label.pack()

    avatar_path = f"{self.username}avatar.jpg"  # 假设头像命名规则
    if os.path.exists(avatar_path):
        image = Image.open(avatar_path)
        # 调整图片为100x100像素的正方形
        image = image.resize((100, 100), Image.ANTIALIAS)
        # 裁剪图片为圆形
        image = self.crop_to_circle(image)
        # 转换为Tkinter的PhotoImage
        photo = ImageTk.PhotoImage(image)
        # 创建Label并放置在正上方
        avatar_label = tk.Label(info_window, image=photo)
        avatar_label.image = photo  # keep a reference!
        avatar_label.pack(side="top")  # 通过side='top'参数放置在正上方

def read_user_info(self):
    filename = "UserInfo.xlsx"
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == self.username:
                self.user_info = {
                    'birthday': row[1].value,
                    'phone': row[2].value,
                    'address': row[3].value
                }
                break

def crop_to_circle(self, image):
    mask = Image.new('L', image.size, 0)
    draw = ImageDraw.Draw(mask)
    draw.ellipse((0, 0) + image.size, fill=255)
    result = Image.new('RGB', image.size, (0, 0, 0))
    result.paste(image, mask=mask)
    return result
