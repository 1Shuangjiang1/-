
from openpyxl import Workbook
from ContactBook_class import *
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook

#这个文件是杨启帆部分的excel写入和读入

def export_to_excel(self):
    # 创建一个Excel工作簿
    workbook = Workbook()
    sheet = workbook.active

    # 添加表头
    headers = ["类别", "姓名", "生日", "电话", "电子邮件", "额外信息"]
    sheet.append(headers)

    # 类别映射到中文
    class_to_chinese = {
        'Classmate': '同学',
        'Teacher': '老师',
        'Colleague': '同事',
        'Friend': '朋友',
        'Relative': '亲戚',
    }
    # 遍历联系人，添加到表格中
    for contact in self.contacts:
        contact_type = class_to_chinese[contact.__class__.__name__]  # 将类名转换为中文类别名
        extra_info = ''
        # 根据类别添加额外信息
        if isinstance(contact, Classmate):
            extra_info = f"{contact.college}; {contact.major}"
        elif isinstance(contact, Teacher):
            extra_info = f"{contact.college}; {contact.title}; {contact.ResearchDirection}"
        elif isinstance(contact, Colleague):
            extra_info = f"{contact.company}"
        elif isinstance(contact, Friend):
            extra_info = f"{contact.how_met}"
        elif isinstance(contact, Relative):
            extra_info = f"{contact.relationship}"

        # 将每个联系人的信息作为一行添加到表格中
        sheet.append(
            [contact_type, contact.name, contact.birthday, contact.phone_number, contact.email, extra_info])

    # 保存工作簿到文件
    # workbook.save(filename="Contacts.xlsx")
    if self.username:  # 确保用户名已设置
        workbook.save(filename=f"{self.username}_Contacts.xlsx")
    else:
        messagebox.showerror("保存失败", "未能识别用户名，保存失败！")


def import_from_excel(self, file_path=None):
    if file_path is None:
        # 弹出文件选择窗口
        file_path = filedialog.askopenfilename(
            title="选择联系人Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
    # 如果用户取消了选择或文件不存在，则不执行任何操作
    if not file_path or not os.path.exists(file_path):
        return

    # 打开工作簿和选择工作表
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 从第二行开始遍历（假设第一行是标题行）
    for row in sheet.iter_rows(min_row=2):
        contact_type, name, birthday, phone_number, email, extra_info = [cell.value for cell in row]

        # 根据提取的额外信息创建联系人对象
        if contact_type == '同学':
            college, major = extra_info.split("; ")
            new_contact = Classmate(name, birthday, phone_number, email, college, major)
        elif contact_type == '老师':
            college, title, research_direction = extra_info.split("; ")
            new_contact = Teacher(name, birthday, phone_number, email, college, title, research_direction)
        elif contact_type == '同事':
            company = extra_info
            new_contact = Colleague(name, birthday, phone_number, email, company)
        elif contact_type == '亲戚':
            relationship = extra_info  # 对于亲戚，额外信息中只有亲戚关系的描述
            new_contact = Relative(name, birthday, phone_number, email, relationship)
        elif contact_type == '朋友':
            how_met = extra_info
            new_contact = Friend(name, birthday, phone_number, email, how_met)
        if new_contact:
            self.contacts.append(new_contact)

    # 更新界面显示或其他操作
    self.update_info_text()


def create_contact_from_excel(self, contact_type, name, birthday, phone_number, email, extra_info):
    # 根据类别和额外信息解析和创建相应的联系人对象
    if contact_type == 'Classmate':
        college, major = extra_info.split("; ")
        return Classmate(name, birthday, phone_number, email, college, major)
    elif contact_type == 'Teacher':
        college, title, research_direction = extra_info.split("; ")
        return Teacher(name, birthday, phone_number, email, college, title, research_direction)
    elif contact_type == 'Colleague':
        company = extra_info
        return Colleague(name, birthday, phone_number, email, company)
    elif contact_type == 'Friend':
        how_met = extra_info
        return Friend(name, birthday, phone_number, email, how_met)
    elif contact_type == 'Relative':
        relationship = extra_info
        return Relative(name, birthday, phone_number, email, relationship)
    # ...其他类别...
    return None