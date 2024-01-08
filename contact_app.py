from ContactBook_class import *
import tkinter as tk
from tkinter import ttk
from xpinyin import Pinyin
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import ttk, filedialog
import matplotlib.pyplot as plt
from collections import Counter
from matplotlib.font_manager import FontProperties
import os
from tkinter import ttk, filedialog, messagebox
import openpyxl
import mail
import random
import string
import datetime

plt.rcParams['font.sans-serif'] = ['SimHei']  # Windows系统使用SimHei字体
plt.rcParams['axes.unicode_minus'] = False  # 正确显示负号



def write_user_to_file(user, filename):
    with open(filename, 'a') as file:
        file.write(f"{user.name},{user.phone_number},{user.__class__.__name__}\n")

def read_users_from_file(filename):
    users = []
    with open(filename, 'r') as file:
        for line in file:
            name, phone_number, user_type = line.strip().split(',')
            if user_type == 'Classmate':
                user = Classmate(name,1 ,phone_number,1,1,1)
            elif user_type == 'Teacher':
                user = Teacher(name, 1,phone_number,1,1,1,1)
            elif user_type == 'Colleague':
                user = Colleague(name,1, phone_number,1,1)
            elif user_type == 'Friend':
                user = Friend(name, 1,phone_number,1,1)
            elif user_type == 'Relative':
                user = Relative(name, 1,phone_number,1,1)
            users.append(user)
    return users

def delete_user_from_file(username, filename):
    # Read all users from the file
    users = read_users_from_file(filename)

    # Find the user with the specified name
    users = [user for user in users if user.name != username]

    # Rewrite the updated user list to the file
    with open(filename, 'w') as file:
        for user in users:
            file.write(f"{user.name},{user.phone_number},{user.__class__.__name__}\n")

def modify_data(filename, target_name, new_data):
    with open(filename, 'r') as file:
        lines = file.readlines()
        for i, line in enumerate(lines):
            # Assuming the name is in the first column
            if line.split(',')[0].strip() == target_name:
                # Modify the line with new data
                lines[i] = f"{new_data}\n"

        with open(filename, 'w') as file:
            file.writelines(lines)
class ContactApp:
    def __init__(self, root):
        self.root = root
        self.root.title("通讯录")
        self.contacts = []
        self.create_gui()
        self.user_file = "users.xlsx"
        self.blacklist_file = "blacklist.xlsx"
        self.username = None
        self.auto_import_enabled = True
        self.email = None  # 邮箱属性
        self.latest_forget_verification_code = None  # 用于保存最新的验证码



    # def login(self, username, password):
    #     if self.check_blacklist(username):
    #         messagebox.showerror("登录失败", "该用户已被列入黑名单，无法登录！")
    #         return False
    #     if not self.check_user_exists(username):
    #         messagebox.showerror("登录失败", "用户名不存在！")
    #         return False
    #     if self.check_password(username, password):
    #         messagebox.showinfo("登录成功", f"欢迎回来，{username}!")
    #         self.username = username  # 保存用户名
    #         if self.auto_import_enabled:
    #             self.auto_import_contacts()  # 自动导入联系人
    #         return True  # 登录成功
    #
    #     else:
    #         messagebox.showerror("登录失败", "密码错误！")
    #         self.add_to_blacklist(username)
    #         messagebox.showinfo("提示", "由于连续登录失败，该用户已被加入黑名单。")
    #         return False  # 登录失败

    def login(self, username, password):
        if self.check_blacklist(username):
            messagebox.showerror("登录失败", "该用户已被列入黑名单，无法登录！")
            return False
        if not self.check_user_exists(username):
            messagebox.showerror("登录失败", "用户名不存在！")
            return False
        if self.check_password(username, password):
            messagebox.showinfo("登录成功", f"欢迎回来，{username}!")
            self.username = username  # 保存用户名
            # 邮箱已在 check_password 方法中设置
            messagebox.showinfo("登录信息", f"用户名: {self.username}, 邮箱: {self.email}")
            if self.auto_import_enabled:
                self.auto_import_contacts()  # 自动导入联系人
            return True  # 登录成功
        else:
            messagebox.showerror("登录失败", "密码错误！")
            self.add_to_blacklist(username)
            messagebox.showinfo("提示", "由于连续登录失败，该用户已被加入黑名单。")
            return False  # 登录失败

    # def login(self):
    #     # 获取用户名和密码（这里应该连接到您的认证系统进行检查）
    #     username = self.username_entry.get()
    #     password = self.password_entry.get()
    #
    #     # 这里的代码只是为了示例，实际上您应该检查用户名和密码是否正确
    #     if self.check_blacklist(username):
    #         messagebox.showerror("登录失败", "该用户已被列入黑名单，无法登录！")
    #     elif not self.check_user_exists(username):
    #         messagebox.showerror("登录失败", "用户名不存在！")
    #     elif self.check_password(username, password):
    #         messagebox.showinfo("登录成功", f"欢迎回来，{username}!")
    #         self.login_window.destroy()  # 关闭登录窗口
    #         self.root.deiconify()  # 显示主窗口
    #     else:
    #         messagebox.showerror("登录失败", "密码错误！")
    #         self.add_to_blacklist(username)
    #         messagebox.showinfo("提示", "由于连续登录失败，该用户已被加入黑名单。")

    # def init_user_excel(self):
    #     if not os.path.exists(self.user_file):
    #         wb = Workbook()
    #         ws = wb.active
    #         ws.append(["Username", "Password"])
    #         wb.save(self.user_file)
    #     else:
    #         wb = load_workbook(self.user_file)
    #     return wb
    def init_user_excel(self):
        if not os.path.exists(self.user_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["用户名", "密码", "邮箱"])  # 添加电子邮件列
            wb.save(self.user_file)
        else:
            wb = load_workbook(self.user_file)
        return wb


    def init_blacklist_excel(self):
        if not os.path.exists(self.blacklist_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Username"])
            wb.save(self.blacklist_file)
        else:
            wb = load_workbook(self.blacklist_file)
        return wb

    def check_user_exists(self, username):
        wb = self.init_user_excel()
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if username == row[0].value:
                return True
        return False

    # def register(self, username, password):
    #     wb = self.init_user_excel()
    #     ws = wb.active
    #     ws.append([username, password])
    #     wb.save(self.user_file)
    #
    # def check_password(self, username, password):
    #     wb = self.init_user_excel()
    #     ws = wb.active
    #     for row in ws.iter_rows(min_row=2):
    #         if username == row[0].value and password == row[1].value:
    #             return True
    #     return False

    def register(self, username, password, email):  # 添加电子邮件参数
        wb = self.init_user_excel()
        ws = wb.active
        ws.append([username, password, email])
        wb.save(self.user_file)

    def check_password(self, username, password):
        wb = self.init_user_excel()
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if username == row[0].value and password == row[1].value:
                # 设置电子邮件属性
                self.email = row[2].value
                return True
        return False

    def add_to_blacklist(self, username):
        wb = self.init_blacklist_excel()
        ws = wb.active
        ws.append([username])
        wb.save(self.blacklist_file)

    def check_blacklist(self, username):
        wb = self.init_blacklist_excel()
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if username == row[0].value:
                return True
        return False

    def create_gui(self):
        # 创建增加联系人按钮
        add_button = ttk.Button(self.root, text="增加联系人", command=self.add_contact)
        add_button.pack()

        # 创建删除联系人按钮
        del_button = ttk.Button(self.root, text="删除联系人", command=self.del_contact)
        del_button.pack()

        mod_button = ttk.Button(self.root, text="修改联系人信息", command=self.mod_contact)
        mod_button.pack()

        sort_button = ttk.Button(self.root, text="按生日排序", command=self.sort_contacts_by_birthday)
        sort_button.pack()

        display_button = ttk.Button(self.root, text="显示联系人", command=self.display_contacts)
        display_button.pack()
        sort_by_surname_button = ttk.Button(self.root, text="按姓氏排序", command=self.sort_contacts_by_surname)
        sort_by_surname_button.pack()

        stats_button = ttk.Button(self.root, text="显示统计图表", command=self.show_charts)
        stats_button.pack()

        exit_button = ttk.Button(self.root, text="退出并保存", command=self.exit_app)
        exit_button.pack()

        # root = tk.Tk()
        # root.title("通讯录")
        #
        # # 创建 Text 组件，设置字体
        # font = ("微软雅黑", 12)
        # info_text = tk.Text(root, height=10, width=30, font=font)
        # info_text.pack()

        table_button = ttk.Button(self.root, text="显示联系人表格", command=self.display_table)
        table_button.pack()

        export_button = ttk.Button(self.root, text="导出到Excel", command=self.export_to_excel)
        export_button.pack()

        import_button = ttk.Button(self.root, text="导入联系人", command=self.import_from_excel)
        import_button.pack()

        simple_search_button = ttk.Button(self.root, text="简化查询", command=self.create_search_gui)
        simple_search_button.pack()

        self.auto_import_button = ttk.Checkbutton(self.root, text="启用自动导入联系人",command=self.toggle_auto_import)
        self.auto_import_button.pack()

        #创建信息框来显示联系人姓名
        self.info_text = tk.Text(self.root, height=10, width=30)
        self.info_text.pack()

    def add_contact(self):
        # 创建增加联系人的窗口
        add_window = tk.Toplevel(self.root)
        add_window.title("增加联系人")
        add_window.geometry("600x400")

        # 下拉框选择联系人类型
        contact_types = ["学生", "同事", "老师", "朋友", "亲戚"]
        type_label = ttk.Label(add_window, text="选择联系人类型:")
        type_label.pack()
        contact_type = ttk.Combobox(add_window, values=contact_types)
        contact_type.pack()

        # 姓名、生日、电话号码、邮件地址输入框
        name_label = ttk.Label(add_window, text="姓名:")
        name_label.pack()
        name_entry = ttk.Entry(add_window)
        name_entry.pack()

        birthday_label = ttk.Label(add_window, text="生日:（输入生日请按照YYYY-MM-DD）")
        birthday_label.pack()
        birthday_entry = ttk.Entry(add_window)
        birthday_entry.pack()

        phone_label = ttk.Label(add_window, text="电话号码:")
        phone_label.pack()
        phone_entry = ttk.Entry(add_window)
        phone_entry.pack()

        email_label = ttk.Label(add_window, text="邮件地址:")
        email_label.pack()
        email_entry = ttk.Entry(add_window)
        email_entry.pack()

        # 根据联系人类型显示特殊信息输入框
        def show_extra_fields():
            selected_type = contact_type.get()
            if selected_type == "学生":
                college_label = ttk.Label(add_window, text="学院:")
                college_label.pack()
                college_entry = ttk.Entry(add_window)
                college_entry.pack()

                major_label = ttk.Label(add_window, text="专业:")
                major_label.pack()
                major_entry = ttk.Entry(add_window)
                major_entry.pack()

            if selected_type == "老师":
                college_label1 = ttk.Label(add_window, text="学院:")
                college_label1.pack()
                college_entry1 = ttk.Entry(add_window)
                college_entry1.pack()

                title_label = ttk.Label(add_window, text="职称:")
                title_label.pack()
                title_entry = ttk.Entry(add_window)
                title_entry.pack()

                ResearchDirection_label = ttk.Label(add_window, text="研究方向:")
                ResearchDirection_label.pack()
                ResearchDirection_entry = ttk.Entry(add_window)
                ResearchDirection_entry.pack()

            if selected_type == "同事":
                company_label = ttk.Label(add_window, text="公司:")
                company_label.pack()
                company_entry = ttk.Entry(add_window)
                company_entry.pack()

            if selected_type == "朋友":
                howmet_label = ttk.Label(add_window, text="如何认识:")
                howmet_label.pack()
                howmet_entry = ttk.Entry(add_window)
                howmet_entry.pack()

            if selected_type == "亲戚":
                relationship_label = ttk.Label(add_window, text="亲戚关系:")
                relationship_label.pack()
                relationship_entry = ttk.Entry(add_window)
                relationship_entry.pack()


            def create_contact():
                # 创建联系人对象并添加到通讯录列表
                selected_type = contact_type.get()

                name = name_entry.get()
                birthday = birthday_entry.get()
                phone_number = phone_entry.get()
                email = email_entry.get()

                if selected_type == "学生":
                    college = college_entry.get()
                    major = major_entry.get()
                    contact = Classmate(name, birthday, phone_number, email, college, major)

                if selected_type == "老师":
                    college = college_entry1.get()
                    title = title_entry.get()
                    ResearchDirection = ResearchDirection_entry.get()
                    contact = Teacher(name, birthday, phone_number, email, college, title, ResearchDirection)

                if selected_type == "同事":
                    company = company_entry.get()
                    contact = Colleague(name, birthday, phone_number, email, company)


                if selected_type == "朋友":
                    howmet = howmet_entry.get()
                    contact = Friend(name, birthday, phone_number, email, howmet)


                if selected_type == "亲戚":
                    relationship = relationship_entry.get()
                    contact = Relative(name, birthday, phone_number, email, relationship)

                self.contacts.append(contact)
                write_user_to_file(contact,"contacts.txt")
                add_window.destroy()
                self.update_info_text()

            create_button = ttk.Button(add_window, text="创建联系人", command=create_contact)
            create_button.pack()

        # 创建联系人按钮
        create_button = ttk.Button(add_window, text="其他细节", command=show_extra_fields)
        create_button.pack()

    def del_contact(self):
        del_window = tk.Toplevel(self.root)
        del_window.title("删除联系人")
        del_window.geometry("600x400")

        contact_names = [contact.name for contact in self.contacts]
        type_label = ttk.Label(del_window, text="选择删除的联系人:")
        type_label.pack()
        contact_type = ttk.Combobox(del_window, values=contact_names)
        contact_type.pack()

        def delete_contact():
            selected_contact = contact_type.get()
            delete_user_from_file(selected_contact, "contacts.txt")
            for contact in self.contacts:
                if contact.name == selected_contact:
                    self.contacts.remove(contact)
                    break
            del_window.destroy()
            self.update_info_text()

        del_button = ttk.Button(del_window, text="确认删除", command=delete_contact)
        del_button.pack()


    def mod_contact(self):
        mod_window = tk.Toplevel(self.root)
        mod_window.title("修改联系人信息")
        mod_window.geometry("600x400")

        mod_names = [contact.name for contact in self.contacts]
        type_label = ttk.Label(mod_window, text="选择要修改信息的联系人:")
        type_label.pack()
        contact_type = ttk.Combobox(mod_window, values=mod_names)
        contact_type.pack()

        type_label = ttk.Label(mod_window, text="请输入要修改的信息:")
        type_label.pack()

        name_label = ttk.Label(mod_window, text="姓名:")
        name_label.pack()
        name_entry = ttk.Entry(mod_window)
        name_entry.pack()

        birthday_label = ttk.Label(mod_window, text="生日:")
        birthday_label.pack()
        birthday_entry = ttk.Entry(mod_window)
        birthday_entry.pack()

        phone_label = ttk.Label(mod_window, text="电话号码:")
        phone_label.pack()
        phone_entry = ttk.Entry(mod_window)
        phone_entry.pack()

        email_label = ttk.Label(mod_window, text="邮件地址:")
        email_label.pack()
        email_entry = ttk.Entry(mod_window)
        email_entry.pack()

        def modify_contact():
            selected_contact = contact_type.get()
            selected_contact = self.contacts[mod_names.index(selected_contact)]
            if name_entry.get() != "":
                initial_name = selected_contact.name
                selected_contact.name = name_entry.get()
            if birthday_entry.get() != "":
                selected_contact.birthday = birthday_entry.get()

            if phone_entry.get() != "":
                selected_contact.phone_number = phone_entry.get()

            if email_entry.get() != "":
                selected_contact.email = email_entry.get()
            print(type(selected_contact.name))
            modify_data("contacts.txt", initial_name,f"{name_entry.get()},{phone_entry.get()},{selected_contact.__class__.__name__}")
            mod_window.destroy()
            self.update_info_text()

        modi_button = ttk.Button(mod_window, text="确认修改", command=modify_contact)
        modi_button.pack()

    def display_contacts(self):
        display_window = tk.Toplevel(self.root)
        display_window.title("显示联系人")
        display_window.geometry("600x400")

        # 下拉框选择类别
        categories = ["全部", "Classmate", "Teacher", "Colleague", "Friend", "Relative"]
        category_label = ttk.Label(display_window, text="选择类别:")
        category_label.pack()
        category_combobox = ttk.Combobox(display_window, values=categories)
        category_combobox.pack()

        # 输入框选择首字母
        letter_label = ttk.Label(display_window, text="输入首字母:")
        letter_label.pack()
        letter_entry = ttk.Entry(display_window)
        letter_entry.pack()

        # 滑动框选择生日区间
        birthday_label = ttk.Label(display_window, text="选择生日区间(按照YYYY-MM-DD输入，可以仅输入年份或月份)")
        birthday_label.pack()

        start_date_var = tk.StringVar()
        end_date_var = tk.StringVar()

        start_date_label = ttk.Label(display_window, text="开始日期:")
        start_date_label.pack()
        start_date_entry = ttk.Entry(display_window, textvariable=start_date_var)
        start_date_entry.pack()

        end_date_label = ttk.Label(display_window, text="结束日期:")
        end_date_label.pack()
        end_date_entry = ttk.Entry(display_window, textvariable=end_date_var)
        end_date_entry.pack()

        def filter_contacts():
            selected_category = category_combobox.get()
            selected_letter = letter_entry.get().lower()
            start_date = start_date_var.get()
            end_date = end_date_var.get()

            filtered_contacts = []

            for contact in self.contacts:
                if selected_category == "全部" or contact.__class__.__name__ == selected_category or not selected_letter:
                    p = Pinyin()
                    if any('\u4e00' <= char <= '\u9fff' for char in contact.name):
                        # 如果包含中文字符，使用xpinyin获取拼音
                        pinyin_name = p.get_pinyin(contact.name, '')
                    else:
                        pinyin_name = contact.name
                    if selected_letter == "" or pinyin_name.lower().startswith(selected_letter):
                        if not start_date and not end_date:
                            filtered_contacts.append(contact)
                        elif not start_date:
                            if contact.birthday <= end_date:
                                filtered_contacts.append(contact)
                        elif not end_date:
                            if contact.birthday >= start_date:
                                filtered_contacts.append(contact)
                        else:
                            if start_date <= contact.birthday <= end_date:
                                filtered_contacts.append(contact)

            self.update_info_text(filtered_contacts)
            display_window.destroy()
        filter_button = ttk.Button(display_window, text="筛选", command=filter_contacts)
        filter_button.pack()



    def update_info_text(self):
        # 清空信息框并更新显示通讯录中的联系人信息
        self.info_text.delete(1.0, tk.END)
        for contact in self.contacts:
            contact_info = str(contact)
            print(contact_info)
            # 使用变量连接的方式来构建要显示的字符串
            contact_info = (
                f"Name: {contact.name}\n"
                f"Birthday: {contact.birthday}\n"
                f"Phone: {contact.phone_number}\n"
                f"Email: {contact.email}\n"
            )

            # 根据联系人类型添加额外的信息
            if isinstance(contact, Classmate):
                contact_info += f"College: {contact.college}\nMajor: {contact.major}\n"
            elif isinstance(contact, Teacher):
                contact_info += f"College: {contact.college}\nTitle: {contact.title}\nResearch Direction: {contact.ResearchDirection}\n"
            elif isinstance(contact, Colleague):
                contact_info += f"Company: {contact.company}\n"
            elif isinstance(contact, Friend):
                contact_info += f"How Met: {contact.how_met}\n"
            elif isinstance(contact, Relative):
                contact_info += f"Relationship: {contact.relationship}\n"

            contact_info += "\n"  # 在每个联系人信息的最后添加一个额外的换行符来分隔

            self.info_text.insert(tk.END, contact_info)  # 将信息插入到文本框

    def sort_contacts_by_birthday(self):
        # 对联系人按生日排序，假设生日格式为YYYY-MM-DD
        self.contacts.sort(key=lambda x: x.birthday)
        self.update_info_text()  # 更新信息框以显示排序后的联系人

    def sort_contacts_by_surname(self):
        # 对联系人按照姓名的拼音进行排序
        p = Pinyin()
        self.contacts.sort(key=lambda x: p.get_pinyin(x.name, ''))
        self.update_info_text()  # 更新信息框以显示排序后的联系人

# _______________________________________________________________________
# 展示联系人信息表格（功能有点重复）
    def display_table(self):
        # 创建一个新窗口来显示表格
        table_window = tk.Toplevel(self.root)
        table_window.title("联系人表格")

        # 创建一个Treeview控件
        columns = ("name", "birthday", "phone", "email", "extra")
        table = ttk.Treeview(table_window, columns=columns, show='headings')

        # 定义列标题
        table.heading("name", text="姓名")
        table.heading("birthday", text="生日")
        table.heading("phone", text="电话")
        table.heading("email", text="电子邮件")
        table.heading("extra", text="额外信息")

        # 添加联系人到表格中
        for contact in self.contacts:
            extra_info = ''
            if isinstance(contact, Classmate):
                extra_info = f"学院: {contact.college}, 专业: {contact.major}"
            elif isinstance(contact, Teacher):
                extra_info = f"学院: {contact.college}, 职称: {contact.title}, 研究方向: {contact.ResearchDirection}"
            elif isinstance(contact, Colleague):
                extra_info = f"公司: {contact.company}"
            elif isinstance(contact, Friend):
                extra_info = f"如何认识: {contact.how_met}"
            elif isinstance(contact, Relative):
                extra_info = f"亲戚关系: {contact.relationship}"

            # 将联系人信息插入表格
            table.insert("", tk.END, values=(contact.name, contact.birthday, contact.phone_number, contact.email, extra_info))

        # 放置表格
        table.pack(expand=True, fill='both')

        # 启动窗口循环
        table_window.mainloop()

# _______________________________________________________________________
# 导出为excel
    def export_to_excel(self):
        # 创建一个Excel工作簿
        workbook = Workbook()
        sheet = workbook.active

        # 添加表头
        headers = ["类别", "姓名", "生日", "电话", "电子邮件", "额外信息"]
        sheet.append(headers)

        # seen = set()
        # unique_contacts = [c for c in self.contacts if (c.name, c.phone_number, c.email) not in seen and not seen.add(
        #     (c.name, c.phone_number, c.email))]

        # 类别映射到中文
        class_to_chinese = {
            'Classmate': '同学',
            'Teacher': '老师',
            'Colleague': '同事',
            'Friend': '朋友',
            'Relative': '亲戚',
        }

        # 遍历联系人，添加到表格中
        for contact in self.contacts:
            contact_type = class_to_chinese[contact.__class__.__name__]  # 将类名转换为中文类别名
            extra_info = ''
            # 根据类别添加额外信息
            if isinstance(contact, Classmate):
                extra_info = f"{contact.college}; {contact.major}"
            elif isinstance(contact, Teacher):
                extra_info = f"{contact.college}; {contact.title}; {contact.ResearchDirection}"
            elif isinstance(contact, Colleague):
                extra_info = f"{contact.company}"
            elif isinstance(contact, Friend):
                extra_info = f"{contact.how_met}"
            elif isinstance(contact, Relative):
                extra_info = f"{contact.relationship}"

            # 将每个联系人的信息作为一行添加到表格中
            sheet.append(
                [contact_type, contact.name, contact.birthday, contact.phone_number, contact.email, extra_info])

        # 保存工作簿到文件
        # workbook.save(filename="Contacts.xlsx")
        if self.username:  # 确保用户名已设置
            workbook.save(filename=f"{self.username}_Contacts.xlsx")
        else:
            messagebox.showerror("保存失败", "未能识别用户名，保存失败！")

# _______________________________________________________________________
# 外部导入excel，读取excel中的联系人

   # def import_from_excel(self):
    #     # 弹出文件选择窗口
    #     excel_path = filedialog.askopenfilename(
    #         title="选择联系人Excel文件",
    #         filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
    #     )
    #     # 如果用户取消了选择，则不执行任何操作
    #     if not excel_path:
    #         return
    def import_from_excel(self, file_path=None):
        if file_path is None:
            # 弹出文件选择窗口
            file_path = filedialog.askopenfilename(
                title="选择联系人Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        # 如果用户取消了选择或文件不存在，则不执行任何操作
        if not file_path or not os.path.exists(file_path):
            return

        # 打开工作簿和选择工作表
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 清除当前所有联系人
        # self.contacts.clear()

        # 从第二行开始遍历（假设第一行是标题行）
        for row in sheet.iter_rows(min_row=2):
            contact_type, name, birthday, phone_number, email, extra_info = [cell.value for cell in row]

            # 根据提取的额外信息创建联系人对象
            if contact_type == '同学':
                college, major = extra_info.split("; ")
                new_contact = Classmate(name, birthday, phone_number, email, college, major)
            elif contact_type == '老师':
                college, title, research_direction = extra_info.split("; ")
                new_contact = Teacher(name, birthday, phone_number, email, college, title, research_direction)
            elif contact_type == '同事':
                company = extra_info
                new_contact = Colleague(name, birthday, phone_number, email, company)
            elif contact_type == '亲戚':
                relationship = extra_info  # 对于亲戚，额外信息中只有亲戚关系的描述
                new_contact = Relative(name, birthday, phone_number, email, relationship)
            elif contact_type == '朋友':
                how_met = extra_info
                new_contact = Friend(name, birthday, phone_number, email, how_met)
            if new_contact:
                self.contacts.append(new_contact)

        # 更新界面显示或其他操作
        self.update_info_text()

    def create_contact_from_excel(self, contact_type, name, birthday, phone_number, email, extra_info):
        # 根据类别和额外信息解析和创建相应的联系人对象
        if contact_type == 'Classmate':
            college, major = extra_info.split("; ")
            return Classmate(name, birthday, phone_number, email, college, major)
        elif contact_type == 'Teacher':
            college, title, research_direction = extra_info.split("; ")
            return Teacher(name, birthday, phone_number, email, college, title, research_direction)
        elif contact_type == 'Colleague':
            company = extra_info
            return Colleague(name, birthday, phone_number, email, company)
        elif contact_type == 'Friend':
            how_met = extra_info
            return Friend(name, birthday, phone_number, email, how_met)
        elif contact_type == 'Relative':
            relationship = extra_info
            return Relative(name, birthday, phone_number, email, relationship)
        # ...其他类别...
        return None

# _______________________________________________________________________
# 简单查询
    def create_search_gui(self):
        # 创建一个新窗口来进行查询
        search_window = tk.Toplevel(self.root)
        search_window.title("简化查询")

        # 创建查询条件的输入框
        self.name_entry = self.create_search_entry(search_window, '姓名:')
        self.birthday_entry = self.create_search_entry(search_window, '生日:')
        self.phone_entry = self.create_search_entry(search_window, '电话:')
        self.email_entry = self.create_search_entry(search_window, '电子邮件:')

        # 创建查询按钮
        search_button = ttk.Button(search_window, text="查询", command=self.simple_search)
        search_button.pack()

    def create_search_entry(self, parent, label_text):
        label = ttk.Label(parent, text=label_text)
        label.pack()
        entry = ttk.Entry(parent)
        entry.pack()
        return entry

    def simple_search(self):
        name = self.name_entry.get().strip().lower()
        birthday = self.birthday_entry.get().strip().lower()
        phone = self.phone_entry.get().strip().lower()
        email = self.email_entry.get().strip().lower()

        # 过滤匹配的联系人
        results = [contact for contact in self.contacts if
                   (not name or name in contact.name.lower()) and
                   (not birthday or birthday in contact.birthday.lower()) and
                   (not phone or phone in contact.phone_number.lower()) and
                   (not email or email in contact.email.lower())]

        # 在这里处理查询结果，例如展示在界面上等
        self.display_search_results(results)

    def display_search_results(self, results):
        # 清空之前的查询结果
        self.info_text.delete('1.0', tk.END)

        # 将查询结果显示在信息框中
        for contact in results:
            self.info_text.insert(tk.END,
                                  f"{contact.name}, {contact.birthday}, {contact.phone_number}, {contact.email}\n")

# _______________________________________________________________________
# 画统计图

    def show_charts(self):
        # 统计各个类型的联系人数量
        contact_types = [contact.__class__.__name__ for contact in self.contacts]
        contact_counts = Counter(contact_types)

        # 生成饼图
        plt.figure(figsize=(8, 4))
        plt.subplot(1, 2, 1)
        plt.pie(contact_counts.values(), labels=contact_counts.keys(), autopct='%1.1f%%')
        plt.title('联系人类型占比')

        # 生成条形图
        plt.subplot(1, 2, 2)
        plt.bar(contact_counts.keys(), contact_counts.values())
        plt.title('联系人类型数量')
        plt.xticks(rotation=45)

        # 显示图表
        plt.tight_layout()
        plt.show()

# _______________________________________________________________________
# 退出自动保存
    def exit_app(self):
        if messagebox.askokcancel("退出", "您确定要退出并保存联系人吗？"):
            # 调用修改后的export_to_excel方法
            self.export_to_excel()
            self.root.destroy()

    def toggle_auto_import(self):
        self.auto_import_enabled = not self.auto_import_enabled
        if self.auto_import_enabled:
            messagebox.showinfo("自动导入", "自动导入功能已开启。")
        else:
            messagebox.showinfo("自动导入", "自动导入功能已关闭。")

    def auto_import_contacts(self):
        file_path = f"{self.username}_Contacts.xlsx"
        if os.path.exists(file_path):
            try:
                # 打开工作簿和选择工作表
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active


                # self.contacts.clear()
                # 清除当前所有联系人

                for row in sheet.iter_rows(min_row=2):
                    contact_type, name, birthday, phone_number, email, extra_info = [cell.value for cell in row]

                    # 根据提取的额外信息创建联系人对象
                    new_contact = None
                    if contact_type == '同学':
                        college, major = extra_info.split("; ")
                        new_contact = Classmate(name, birthday, phone_number, email, college, major)
                    elif contact_type == '老师':
                        college, title, research_direction = extra_info.split("; ")
                        new_contact = Teacher(name, birthday, phone_number, email, college, title, research_direction)
                    elif contact_type == '同事':
                        company = extra_info
                        new_contact = Colleague(name, birthday, phone_number, email, company)
                    elif contact_type == '亲戚':
                        relationship = extra_info  # 对于亲戚，额外信息中只有亲戚关系的描述
                        new_contact = Relative(name, birthday, phone_number, email, relationship)
                    elif contact_type == '朋友':
                        how_met = extra_info
                        new_contact = Friend(name, birthday, phone_number, email, how_met)

                    if new_contact:
                        self.contacts.append(new_contact)

                self.update_info_text()  # 更新界面显示
                messagebox.showinfo("自动导入", "联系人已成功自动导入。")
                self.check_and_send_birthday_wish()
            except Exception as e:
                messagebox.showerror("自动导入", f"导入失败: {e}")


    # def auto_import_contacts(self):
    #     file_path = f"{self.username}_Contacts.xlsx"
    #     if os.path.exists(file_path):
    #         self.import_from_excel(file_path)  # 调用现有的导入方法
    #     else:
    #         messagebox.showinfo("自动导入", "没有找到之前的联系人文件，自动导入未执行。")

    # def register_user(self, username, password, email):
    #     if self.check_user_exists(username):
    #         messagebox.showerror("注册失败", "该用户名已存在！")
    #         return False
    #     else:
    #         self.add_user(username, password, email)
    #         messagebox.showinfo("注册成功", f"欢迎，{username}! 请使用您的新账号登录。")
    #         return True

# _______________________________________________________________________
    # 找回密码
    def generate_random_code(self,length=6):
        """生成一个指定长度的随机验证码，包含数字和字母。"""
        characters = string.ascii_letters + string.digits  # 包括字母和数字
        return ''.join(random.choice(characters) for i in range(length))

    def send_verification_code(self, email):
        # 在这里实现发送验证码到邮箱的逻辑
        # 为简化示例，这里返回一个假的验证码 "123456"
        # print(f"发送验证码到 {email}")
        # return "123456"
        host_server = 'smtp.qq.com'
        sender_qq = 'yangqifanbq@qq.com'
        pwd = 'ligsaipzxolvhcec'  # 注意这通常是邮箱的授权码，并非登录密码
        print(email)
        receiver = email  # 直接使用方法参数
        mail_title = '邮箱注册验证码'
        # verification_code = '123456'  # 正常情况下这里应该是随机生成的验证码
        verification_code = self.generate_random_code(6)
        mail_content = f'<p>这是使用python登录QQ邮箱发送HTNL格式邮件的测试：您的邮箱验证码为：{verification_code}</p>'


        try:
            # 调用封装好的邮件发送函数
            mail.send_verification_email(host_server, sender_qq, pwd, receiver, mail_title, mail_content)
            return verification_code
        except Exception as e:
            print("邮件发送失败:", e)
            return None

    def validate_code(self, input_code, correct_code):
        # 校验输入的验证码是否与正确的验证码匹配
        return input_code == correct_code

    def register_user(self, username, password, email):
        if self.check_user_exists(username):
            messagebox.showerror("注册失败", "该用户名已存在！")
            return False
        if self.check_email_exists(email):
            messagebox.showerror("注册失败", "该邮箱已被注册！")
            return False
        # 获取用户输入的验证码
        correct_code = self.send_verification_code(email)
        verification_window = tk.Toplevel(self.root)
        verification_window.title("请输入验证码")

        verification_label = ttk.Label(verification_window, text="验证码:")
        verification_label.pack()
        verification_entry = ttk.Entry(verification_window)
        verification_entry.pack()

        def on_verification_confirm():
            input_code = verification_entry.get()
            if self.validate_code(input_code, correct_code):
                self.add_user(username, password, email)
                messagebox.showinfo("注册成功", f"欢迎，{username}! 请使用您的新账号登录。")
                verification_window.destroy()
            else:
                messagebox.showerror("注册失败", "验证码错误！")
                verification_entry.delete(0, tk.END)  # 清空输入框

        confirm_button = ttk.Button(verification_window, text="验证", command=on_verification_confirm)
        confirm_button.pack()

    def check_email_exists(self, email):
        """检查给定邮箱是否已经存在于用户信息中"""
        wb = self.init_user_excel()
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            username_cell, password_cell, email_cell = row
            if email == email_cell.value:
                return True
        return False

    def add_user(self, username, password, email):
        wb = self.init_user_excel()
        ws = wb.active
        ws.append([username, password, email])  # 假设 Excel 文件有三列：用户名、密码和邮箱
        wb.save(self.user_file)

#_______________________________________________________________________
    #找回密码
    def send_reset_code(self, email):
        # # 这里发送重置密码的验证码到用户的邮箱
        # # 并返回生成的验证码，简化示例中返回一个固定验证码
        # reset_code = "123456"
        # # 实际应用中，这里将调用之前定义的邮件发送方法
        # print("发送重置密码验证码到", email)
        # return reset_code
        # 在这里实现发送验证码到邮箱的逻辑
        # 为简化示例，这里返回一个假的验证码 "123456"
        # print(f"发送验证码到 {email}")
        # return "123456"
        host_server = 'smtp.qq.com'
        sender_qq = 'yangqifanbq@qq.com'
        pwd = 'ligsaipzxolvhcec'  # 注意这通常是邮箱的授权码，并非登录密码
        print(email)
        receiver = email  # 直接使用方法参数
        mail_title = '邮箱注册验证码'
        # verification_code = '123456'  # 正常情况下这里应该是随机生成的验证码
        verification_code = self.generate_random_code(6)
        self.latest_forget_verification_code=verification_code
        mail_content = f'<p>这是使用python登录QQ邮箱发送HTNL格式邮件的测试：您的邮箱验证码为：{verification_code}</p>'


        try:
            # 调用封装好的邮件发送函数
            mail.send_verification_email(host_server, sender_qq, pwd, receiver, mail_title, mail_content)
            return verification_code
        except Exception as e:
            print("邮件发送失败:", e)
            return None

    def reset_password(self, email, new_password):
        # 找到与提供的邮箱对应的用户名和原密码，并更新密码
        wb = openpyxl.load_workbook(self.user_file)
        ws = wb.active
        user_found = False

        # 遍历Excel的行
        for row in ws.iter_rows(min_row=2):
            username_cell, password_cell, email_cell = row
            # 检查邮箱是否匹配
            if email_cell.value == email:
                # 更新密码
                password_cell.value = new_password
                user_found = True
                break

        if user_found:
            # 保存工作簿
            wb.save(self.user_file)
            messagebox.showinfo("重置密码", "密码重置成功，请使用新密码登录。")
        else:
            messagebox.showerror("重置密码", "没有找到匹配的邮箱，请确保您输入的邮箱正确。")

        # 关闭工作簿
        wb.close()


    # def check_and_send_birthday_wish(self):
    #     today = datetime.date.today()
    #     tomorrow = today + datetime.timedelta(days=1)
    #
    #     # 只比较月和日
    #     if (self.birthday.month, self.birthday.day) in [(today.month, today.day), (tomorrow.month, tomorrow.day)]:
    #         self.send_birthday_email()
    #
    # def send_birthday_email(self):
    #     # 实现发送邮件的逻辑，这里仅为示例
    #     print(f"Sending birthday email to {self.name} at {self.email}")
    #     # 实际发送邮件的代码应该在这里
    def check_and_send_birthday_wish(self):
        today = datetime.date.today()
        tomorrow = today + datetime.timedelta(days=1)

        for contact in self.contacts:
            # 检查联系人是否有今天或明天的生日
            birthday = datetime.datetime.strptime(contact.birthday, "%Y-%m-%d").date()
            if (birthday.month, birthday.day) in [(today.month, today.day), (tomorrow.month, tomorrow.day)]:
                # 如果是，发送生日祝福邮件
                self.edit_and_send_birthday_email(contact)


    # def edit_and_send_birthday_email(self, contact):
    #     # 创建一个新窗口
    #     edit_window = tk.Toplevel(self.root)
    #     edit_window.title("编辑生日邮件")
    #
    #     # 邮件标题输入
    #     tk.Label(edit_window, text="邮件标题:").pack()
    #     title_entry = tk.Entry(edit_window)
    #     title_entry.insert(0, '生日快乐')  # 默认值
    #     title_entry.pack()
    #
    #     # 邮件内容输入
    #     tk.Label(edit_window, text="邮件内容:").pack()
    #     content_entry = tk.Text(edit_window, height=10)
    #     content_entry.insert('1.0', f'亲爱的{contact.name}，祝您生日快乐！')  # 默认值
    #     content_entry.pack()
    #
    #     # 发送按钮
    #     send_button = tk.Button(edit_window, text="发送邮件",
    #                             command=lambda: self.send_birthday_email(contact,
    #                                                                      title_entry.get(),
    #                                                                      content_entry.get('1.0', 'end')))
    #     send_button.pack()

    def edit_and_send_birthday_email(self, contact):
        # 创建一个新窗口
        edit_window = tk.Toplevel(self.root)
        edit_window.title("编辑生日邮件")

        # 确定生日是今天还是明天
        today = datetime.date.today()
        tomorrow = today + datetime.timedelta(days=1)
        birthday = datetime.datetime.strptime(contact.birthday, "%Y-%m-%d").date()
        birthday_str = "今天" if (birthday.month, birthday.day) == (today.month, today.day) else "明天"

        # 显示是今天还是明天生日
        tk.Label(edit_window, text=f"{contact.name} 的生日是{birthday_str}!").pack()

        # 查看详细信息按钮
        details_button = tk.Button(edit_window, text="查看详细信息", command=lambda: self.show_contact_details(contact))
        details_button.pack()

        # 邮件标题输入
        tk.Label(edit_window, text="邮件标题:").pack()
        title_entry = tk.Entry(edit_window)
        title_entry.insert(0, '生日快乐')  # 默认值
        title_entry.pack()

        # 邮件内容输入
        tk.Label(edit_window, text="邮件内容:").pack()
        content_entry = tk.Text(edit_window, height=10)
        content_entry.insert('1.0', f'亲爱的{contact.name}，祝您生日快乐！')  # 默认值
        content_entry.pack()

        # 发送按钮
        send_button = tk.Button(edit_window, text="发送邮件",
                                command=lambda: self.send_birthday_email(contact,
                                                                         title_entry.get(),
                                                                         content_entry.get('1.0', 'end'),edit_window))
        send_button.pack()

    def show_contact_details(self, contact):
        # 创建一个新窗口显示联系人详细信息
        details_window = tk.Toplevel(self.root)
        details_window.title(f"{contact.name} 的详细信息")

        # 根据不同类型的联系人显示详细信息
        tk.Label(details_window, text=f"姓名: {contact.name}").pack()
        tk.Label(details_window, text=f"生日: {contact.birthday}").pack()
        tk.Label(details_window, text=f"电话: {contact.phone_number}").pack()
        tk.Label(details_window, text=f"电子邮件: {contact.email}").pack()

        # 以下为不同类型联系人特有的信息
        if isinstance(contact, Classmate):
            tk.Label(details_window, text=f"大学: {contact.college}").pack()
            tk.Label(details_window, text=f"专业: {contact.major}").pack()
        elif isinstance(contact, Teacher):
            tk.Label(details_window, text=f"大学: {contact.college}").pack()
            tk.Label(details_window, text=f"职称: {contact.title}").pack()
            tk.Label(details_window, text=f"研究方向: {contact.ResearchDirection}").pack()
        elif isinstance(contact, Colleague):
            tk.Label(details_window, text=f"公司: {contact.company}").pack()
        elif isinstance(contact, Friend):
            tk.Label(details_window, text=f"认识途径: {contact.how_met}").pack()
        elif isinstance(contact, Relative):
            tk.Label(details_window, text=f"关系: {contact.relationship}").pack()
    def send_birthday_email(self, contact, mail_title, mail_content,edit_window):
        # 发送生日祝福邮件的逻辑

        host_server = 'smtp.qq.com'
        sender_qq = 'yangqifanbq@qq.com'
        pwd = 'ligsaipzxolvhcec'  # 注意这通常是邮箱的授权码，并非登录密码
        receiver = contact.email


        mail.send_verification_email(host_server, sender_qq, pwd, receiver, mail_title, mail_content)
        print(f"发送生日祝福邮件给 {contact.name} 标题: {mail_title} 内容: {mail_content}")
        edit_window.destroy()



