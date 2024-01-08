from ContactBook_class import *
import tkinter as tk
from tkinter import simpledialog
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import os
import contact_app

# def inherit_contacts(self):
#     inherit_window = tk.Toplevel(self.root)
#     inherit_window.title("继承联系人")
#
#     # 输入用户名
#     tk.Label(inherit_window, text="用户名:").pack()
#     username_entry = tk.Entry(inherit_window)
#     username_entry.pack()
#
#     # 输入密码
#     tk.Label(inherit_window, text="密码:").pack()
#     password_entry = tk.Entry(inherit_window, show="*")
#     password_entry.pack()
#
#     # 继承按钮
#     inherit_button = tk.Button(inherit_window, text="继承联系人",
#                                command=lambda: self.add_inherited_contacts(
#                                    username_entry.get(),
#                                    password_entry.get(),
#                                    inherit_window))
#     inherit_button.pack()
def inherit_contacts(self):
    # 创建一个新窗口来输入其他账户的用户名和密码
    inherit_window = tk.Toplevel(self.root)
    inherit_window.title("继承联系人")

    tk.Label(inherit_window, text="其他用户名:").pack()
    other_username_entry = tk.Entry(inherit_window)
    other_username_entry.pack()

    tk.Label(inherit_window, text="其他密码:").pack()
    other_password_entry = tk.Entry(inherit_window, show="*")
    other_password_entry.pack()

    confirm_button = ttk.Button(inherit_window, text="确认",
                                command=lambda: self.add_inherited_contacts(
                                    other_username_entry.get(),
                                    other_password_entry.get(),
                                    inherit_window))
    confirm_button.pack()


# def add_inherited_contacts(self, other_username, other_password, window):
#     # 验证另一个账号的用户名和密码
#     if self.validate_other_user(other_username, other_password):
#         other_contacts = self.read_other_user_contacts(other_username)
#         # 添加到当前用户的联系人列表
#         self.contacts.extend(other_contacts)
#         messagebox.showinfo("成功", "联系人已继承！")
#         window.destroy()
#         # contact_app.
#     else:
#         messagebox.showerror("错误", "用户名或密码错误。")




# def validate_other_user(self, other_username, other_password):
#     # 这里应该是验证另一个用户的逻辑
#     # 简单示例，应该是到数据库中检查用户名和密码
#     return True  # 假设验证总是通过，实际应用中应该是真正的验证逻辑
#
#
# def read_other_user_contacts(self, other_username):
#     # 读取另一个用户的联系人信息
#     filename = f"{other_username}_Contacts.xlsx"
#     if not os.path.exists(filename):
#         messagebox.showerror("错误", "联系人文件不存在。")
#         return []
#
#     workbook = load_workbook(filename)
#     sheet = workbook.active
#     other_contacts = []
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         contact_info = {
#             'name': row[0],
#             'birthday': row[1],
#             'phone_number': row[2],
#             'email': row[3],
#             # 假设Excel文件中还有其他列
#         }
#         other_contacts.append(contact_info)
#     return other_contacts